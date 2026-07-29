"""
Excel 导出模块
将未举办的活动导出为 .xlsx 文件。
"""
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import EXPORTS_DIR, EXPORT_FIELDS, TOPIC_LABELS

logger = logging.getLogger(__name__)


def _format_date(event):
    """格式化活动日期为可读字符串。"""
    y = event.get("year", "")
    m = event.get("month", "")
    d = event.get("day", "")
    if y and m and d:
        return f"{y}年{m}月{d}日"
    return "日期待定"


def _format_topics(event):
    """将议题标签列表转为中文逗号分隔字符串。"""
    topics = event.get("topics", [])
    return "、".join(TOPIC_LABELS.get(t, t) for t in topics)


def export_unheld_events(events, output_dir=None):
    """
    将未举办的活动 (status 为 confirmed 或 upcoming) 导出为 Excel。
    返回生成的文件路径。
    """
    output_dir = output_dir or EXPORTS_DIR
    os.makedirs(output_dir, exist_ok=True)

    # 筛选未举办活动
    unheld = [
        e for e in events
        if e.get("status") in ("confirmed", "upcoming")
    ]

    # 按日期排序
    unheld.sort(key=lambda e: (
        e.get("year", 9999),
        e.get("month", 99),
        e.get("day", 99),
    ))

    if not unheld:
        logger.info("没有未举办的活动需要导出")
        # 仍然生成一个空文件，标明本次运行无未举办活动

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "未举办活动"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F81F7", end_color="2F81F7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    # 状态颜色
    status_fills = {
        "confirmed": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "upcoming": PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid"),
    }

    # 写表头
    for col, field in enumerate(EXPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for row_idx, event in enumerate(unheld, 2):
        values = [
            event.get("title", ""),
            _format_date(event),
            event.get("organizer", ""),
            event.get("typeLabel", ""),
            _format_topics(event),
            event.get("statusLabel", ""),
            event.get("venue", "") or "待确认",
            event.get("registration", "") or "详见活动链接",
            event.get("url", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            # 状态列着色
            if col == 6 and event.get("status") in status_fills:
                cell.fill = status_fills[event["status"]]

    # 列宽设置
    col_widths = [30, 16, 25, 10, 20, 10, 25, 25, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 行高
    ws.row_dimensions[1].height = 30
    for r in range(2, len(unheld) + 2):
        ws.row_dimensions[r].height = 25

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    if unheld:
        ws.auto_filter.ref = f"A1:I{len(unheld) + 1}"

    # 生成文件名
    today_str = datetime.now().strftime("%Y%m%d")
    filename = f"未举办活动_{today_str}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # 如果文件已存在，加序号
    counter = 1
    while os.path.exists(filepath):
        filename = f"未举办活动_{today_str}v{counter}.xlsx"
        filepath = os.path.join(output_dir, filename)
        counter += 1

    wb.save(filepath)
    logger.info("Excel 导出完成: %s (%d 条未举办活动)", filepath, len(unheld))
    return filepath
